from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

from openpyxl import Workbook, load_workbook

#Excel
wb = load_workbook('WIEBETECH PRICES.xlsx')
ws = wb.active
max_row = ws.max_row
total = 0

product_codes = []

for row in range(2, max_row + 1):
    product_code = ws[f'A{row}'].value
    if product_code is not None:
        product_codes.append(product_code)

#Web
service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

base_url = 'https://wiebetech.com'

def get_product_link(product_code):
    search_url = f"{base_url}?s={product_code}"
    page = driver.get(search_url)
    find = driver.find_element(By.XPATH, "//*[@id='content']/div/div/div/div[1]/div[1]/a")
    find.click()
    product_url = driver.current_url
    return product_url

row = 1
for code in product_codes:
    row += 1
    try:
        link = get_product_link(code)
        ws[f'I{row}'].value = link
    except:
        ws[f'I{row}'].value = "Product is not available"
        continue

wb.save('WIEBETECH PRICES AND LINKS.xlsx')
wb.close()
